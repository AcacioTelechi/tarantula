import csv
import io
import json

from typer.testing import CliRunner

from tarantula.cli import app
from tarantula.flatten import flatten_run


def _payload():
    return {
        "run_id": 12,
        "started_at": "2026-04-20T17:20:58Z",
        "finished_at": "2026-04-20T17:32:03Z",
        "sites": [
            {
                "seed_url": "https://a.example",
                "crawl_status": "ok",
                "pages_fetched": 800,
                "variables": {
                    "nome": {
                        "value": "Instituto A",
                        "source_url": "https://a.example/sobre",
                        "quote": "O Instituto A faz coisas",
                        "reasoning": "consistente",
                        "required_missing": False,
                    },
                    "ano_fund": {
                        "value": 1998,
                        "source_url": "https://a.example/historia",
                        "quote": "fundado em 1998",
                        "reasoning": "data citada",
                        "required_missing": False,
                    },
                },
            },
            {
                "seed_url": "https://b.example",
                "crawl_status": "partial",
                "pages_fetched": 10,
                "variables": {
                    "nome": {
                        "value": "Instituto B",
                        "source_url": "https://b.example/",
                        "quote": "Instituto B",
                        "reasoning": "homepage",
                        "required_missing": False,
                    },
                    "cnpj": {
                        "value": None,
                        "source_url": None,
                        "quote": None,
                        "reasoning": "not found",
                        "required_missing": True,
                    },
                },
            },
        ],
    }


def test_flatten_run_one_row_per_site_with_run_and_site_columns():
    fields, rows = flatten_run(_payload())

    assert len(rows) == 2
    # Leading run/site columns come first, in this order.
    assert fields[:6] == [
        "run_id", "started_at", "finished_at",
        "seed_url", "crawl_status", "pages_fetched",
    ]
    assert rows[0]["run_id"] == 12
    assert rows[0]["seed_url"] == "https://a.example"
    assert rows[0]["pages_fetched"] == 800
    assert rows[1]["crawl_status"] == "partial"


def test_flatten_run_expands_each_variable_into_five_columns():
    fields, rows = flatten_run(_payload())

    for sub in ("value", "source_url", "quote", "reasoning", "required_missing"):
        assert f"nome_{sub}" in fields
    # Sub-fields appear grouped per variable, in a fixed order.
    i = fields.index("nome_value")
    assert fields[i:i + 5] == [
        "nome_value", "nome_source_url", "nome_quote",
        "nome_reasoning", "nome_required_missing",
    ]
    assert rows[0]["nome_value"] == "Instituto A"
    assert rows[0]["nome_quote"] == "O Instituto A faz coisas"
    assert rows[0]["ano_fund_value"] == 1998


def test_flatten_run_unions_variables_across_sites_first_seen_order():
    fields, rows = flatten_run(_payload())

    # nome and ano_fund first seen on site 0, cnpj first appears on site 1.
    var_cols = [f for f in fields if f.endswith("_value")]
    assert var_cols == ["nome_value", "ano_fund_value", "cnpj_value"]
    # A site missing a variable gets empty cells for it.
    assert rows[0]["cnpj_value"] == ""
    assert rows[1]["ano_fund_value"] == ""
    # None values render as empty cells too.
    assert rows[1]["cnpj_value"] == ""
    assert rows[1]["cnpj_required_missing"] is True


def test_flatten_command_writes_csv_file(tmp_path):
    in_path = tmp_path / "out.json"
    in_path.write_text(json.dumps(_payload()))
    out_path = tmp_path / "out.csv"

    runner = CliRunner()
    result = runner.invoke(app, [
        "flatten", str(in_path), "--output", str(out_path),
    ], catch_exceptions=False)
    assert result.exit_code == 0, result.output

    with out_path.open(newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert len(rows) == 2
    assert rows[0]["nome_value"] == "Instituto A"
    assert rows[0]["seed_url"] == "https://a.example"
    assert rows[1]["cnpj_required_missing"] == "True"


def test_flatten_command_writes_csv_to_stdout_without_output(tmp_path):
    in_path = tmp_path / "out.json"
    in_path.write_text(json.dumps(_payload()))

    runner = CliRunner()
    result = runner.invoke(app, ["flatten", str(in_path)], catch_exceptions=False)
    assert result.exit_code == 0, result.output

    reader = csv.DictReader(io.StringIO(result.stdout))
    rows = list(reader)
    assert len(rows) == 2
    assert rows[0]["nome_value"] == "Instituto A"


def test_write_xlsx_header_and_rows(tmp_path):
    from openpyxl import load_workbook
    from tarantula.flatten import write_xlsx
    fields, rows = flatten_run(_payload())
    out = tmp_path / "out.xlsx"
    write_xlsx(fields, rows, out)

    ws = load_workbook(out).active
    header = [c.value for c in ws[1]]
    assert header == fields
    data = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data) == 2
    # nome_value column matches the payload for the first site.
    nome_col = fields.index("nome_value")
    assert data[0][nome_col] == "Instituto A"


def test_write_xlsx_stringifies_list_values(tmp_path):
    # Array variables put a list in their _value column; xlsx cells must be
    # scalar, so the list is stringified rather than crashing openpyxl.
    from openpyxl import load_workbook
    from tarantula.flatten import write_xlsx
    payload = {
        "run_id": 1, "started_at": "t", "finished_at": "t",
        "sites": [{
            "seed_url": "s", "crawl_status": "ok", "pages_fetched": 1,
            "variables": {
                "emp_fin": {"value": ["Acme", "Globex"], "sources": [],
                            "reasoning": "r", "required_missing": False},
            },
        }],
    }
    fields, rows = flatten_run(payload)
    out = tmp_path / "arr.xlsx"
    write_xlsx(fields, rows, out)
    ws = load_workbook(out).active
    col = fields.index("emp_fin_value")
    cell = list(ws.iter_rows(min_row=2, values_only=True))[0][col]
    assert isinstance(cell, str)
    assert "Acme" in cell and "Globex" in cell


def test_flatten_command_writes_xlsx_by_extension(tmp_path):
    from openpyxl import load_workbook
    in_path = tmp_path / "out.json"
    in_path.write_text(json.dumps(_payload()))
    out_path = tmp_path / "out.xlsx"

    result = CliRunner().invoke(app, [
        "flatten", str(in_path), "-o", str(out_path),
    ], catch_exceptions=False)
    assert result.exit_code == 0, result.output

    ws = load_workbook(out_path).active
    header = [c.value for c in ws[1]]
    assert "nome_value" in header
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
